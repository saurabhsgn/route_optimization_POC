"""FastAPI application — single entry point for the Route Optimizer POC."""

from datetime import datetime, timedelta
from io import BytesIO
from fastapi import FastAPI, Query, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel
from pathlib import Path
import openpyxl

from config import (
    DEFAULT_NUM_ORDERS, DEFAULT_NUM_DRIVERS, DEFAULT_NUM_VEHICLES,
    OPERATING_START_HOUR, STORE_LAT, STORE_LNG, SLA_WINDOW_HOURS,
)
from models import Order, OrderStatus, Priority
from data_generator import generate_orders, generate_drivers, generate_vehicles
from optimizer import RouteOptimizer
from visualizer import generate_route_map, generate_chart_data
from geocoder import geocode_address

app = FastAPI(title="Route Optimizer POC")

# In-memory state
state = {
    "orders": [],
    "drivers": [],
    "vehicles": [],
    "result": None,
    "reference_time": None,
}


# --- Request models ---

class ManualOrder(BaseModel):
    customer_name: str = ""
    address: str
    priority: str = "MEDIUM"
    delivery_datetime: str | None = None
    latitude: float | None = None
    longitude: float | None = None


class ManualOrderBatch(BaseModel):
    orders: list[ManualOrder]
    num_drivers: int = 3
    num_vehicles: int = 3


# --- Endpoints ---

@app.get("/", response_class=HTMLResponse)
async def dashboard():
    html_path = Path(__file__).parent / "templates" / "dashboard.html"
    return HTMLResponse(html_path.read_text(encoding="utf-8"))


@app.post("/api/generate")
async def generate_data(
    num_orders: int = Query(DEFAULT_NUM_ORDERS, ge=3, le=50),
    num_drivers: int = Query(DEFAULT_NUM_DRIVERS, ge=1, le=10),
    num_vehicles: int = Query(DEFAULT_NUM_VEHICLES, ge=1, le=10),
    scenario: str = Query("normal"),
):
    ref_time = datetime.now().replace(
        hour=OPERATING_START_HOUR, minute=0, second=0, microsecond=0
    )
    state["reference_time"] = ref_time
    state["orders"] = generate_orders(num_orders, ref_time, scenario)
    state["drivers"] = generate_drivers(num_drivers, ref_time)
    state["vehicles"] = generate_vehicles(num_vehicles)
    state["result"] = None

    return {
        "orders": [o.model_dump(mode="json") for o in state["orders"]],
        "drivers": [d.model_dump(mode="json") for d in state["drivers"]],
        "vehicles": [v.model_dump(mode="json") for v in state["vehicles"]],
        "count": {
            "orders": len(state["orders"]),
            "drivers": len(state["drivers"]),
            "vehicles": len(state["vehicles"]),
        },
    }


@app.post("/api/geocode")
async def geocode_single(address: str = Query(...)):
    """Geocode a single US address and return lat/lng."""
    lat, lng, display = geocode_address(address)
    return {
        "address": address,
        "latitude": lat,
        "longitude": lng,
        "display_name": display,
        "found": lat is not None,
    }


@app.post("/api/manual-orders")
async def submit_manual_orders(batch: ManualOrderBatch):
    """
    Accept manually entered orders with addresses.
    Geocodes addresses without lat/lng, uses provided coords when available.
    Supports delivery_datetime for per-order SLA deadlines.
    """
    ref_time = datetime.now().replace(
        hour=OPERATING_START_HOUR, minute=0, second=0, microsecond=0
    )
    state["reference_time"] = ref_time
    state["result"] = None

    created_orders = []
    geocode_errors = []

    for i, mo in enumerate(batch.orders):
        # Use provided lat/lng or geocode the address
        if mo.latitude is not None and mo.longitude is not None:
            lat, lng, display = mo.latitude, mo.longitude, mo.address
        else:
            lat, lng, display = geocode_address(mo.address)
            if lat is None:
                geocode_errors.append({
                    "index": i,
                    "address": mo.address,
                    "error": "Address not found — check spelling or add city/state",
                })
                continue

        priority = Priority(mo.priority) if mo.priority in ("HIGH", "MEDIUM", "LOW") else Priority.MEDIUM

        # Use provided delivery_datetime as SLA deadline, or fallback to priority-based
        if mo.delivery_datetime:
            try:
                sla_deadline = datetime.fromisoformat(mo.delivery_datetime)
            except ValueError:
                sla_hours = SLA_WINDOW_HOURS[priority.value]
                sla_deadline = ref_time + timedelta(hours=sla_hours)
        else:
            sla_hours = SLA_WINDOW_HOURS[priority.value]
            sla_deadline = ref_time + timedelta(hours=sla_hours)

        order = Order(
            order_id=f"ORD-{i + 1:04d}",
            priority=priority,
            status=OrderStatus.READY,
            created_at=ref_time,
            ready_at=ref_time + timedelta(minutes=10),
            sla_deadline=sla_deadline,
            customer_name=mo.customer_name or f"Customer {i + 1}",
            address=display or mo.address,
            latitude=round(lat, 6),
            longitude=round(lng, 6),
            service_time_min=5,
            weight_kg=1.0,
            is_peak=False,
            can_defer=priority == Priority.LOW,
            return_trip_eligible=False,
        )
        created_orders.append(order)

    state["orders"] = created_orders
    state["drivers"] = generate_drivers(batch.num_drivers, ref_time)
    state["vehicles"] = generate_vehicles(batch.num_vehicles)

    return {
        "orders": [o.model_dump(mode="json") for o in created_orders],
        "geocode_errors": geocode_errors,
        "count": {
            "orders": len(created_orders),
            "drivers": batch.num_drivers,
            "vehicles": batch.num_vehicles,
            "failed": len(geocode_errors),
        },
    }


@app.get("/api/template")
async def download_template():
    """Download a sample Excel template for order upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = ["customer_name", "address", "priority", "delivery_datetime", "latitude", "longitude"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    sample_rows = [
        ["John Smith", "350 5th Ave, New York, NY 10118", "HIGH", "2026-03-15 14:00", "", ""],
        ["Jane Doe", "1 World Trade Center, New York, NY", "MEDIUM", "2026-03-15 16:00", "", ""],
        ["Bob Wilson", "30 Rockefeller Plaza, New York, NY", "LOW", "", "", ""],
        ["Alice Brown", "200 Central Park West, New York, NY", "HIGH", "2026-03-15 13:30", "", ""],
        ["Tom Davis", "Brooklyn Bridge, Brooklyn, NY", "MEDIUM", "", "", ""],
        ["Pre-geocoded", "Custom Location", "HIGH", "2026-03-15 12:00", "40.7484", "-73.9857"],
    ]
    for row in sample_rows:
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=order_template.xlsx"},
    )


@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Parse an uploaded Excel file and return orders for review."""
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    # Read headers from row 1
    headers = []
    for cell in ws[1]:
        val = str(cell.value or "").strip().lower().replace(" ", "_")
        headers.append(val)

    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        row_dict = {}
        for idx, val in enumerate(row):
            if idx < len(headers):
                row_dict[headers[idx]] = val

        address = str(row_dict.get("address", "") or "").strip()
        if not address:
            continue

        # Parse delivery_datetime (handle Excel datetime objects and strings)
        dt_raw = row_dict.get("delivery_datetime") or row_dict.get("delivery_date") or ""
        delivery_dt = ""
        if isinstance(dt_raw, datetime):
            delivery_dt = dt_raw.isoformat()
        elif dt_raw:
            delivery_dt = str(dt_raw).strip()

        # Parse lat/lng
        lat = None
        lng = None
        try:
            lat_raw = row_dict.get("latitude") or row_dict.get("lat")
            lng_raw = row_dict.get("longitude") or row_dict.get("lng") or row_dict.get("lon")
            if lat_raw and lng_raw:
                lat = float(lat_raw)
                lng = float(lng_raw)
        except (ValueError, TypeError):
            pass

        orders.append({
            "customer_name": str(row_dict.get("customer_name", "") or row_dict.get("name", "") or ""),
            "address": address,
            "priority": str(row_dict.get("priority", "MEDIUM") or "MEDIUM").upper(),
            "delivery_datetime": delivery_dt,
            "latitude": lat,
            "longitude": lng,
        })

    return {"orders": orders, "count": len(orders)}


@app.post("/api/optimize")
async def optimize_routes():
    if not state["orders"]:
        return JSONResponse({"error": "No orders. Generate random data or submit manual orders first."}, 400)

    optimizer = RouteOptimizer(
        orders=state["orders"],
        drivers=state["drivers"],
        vehicles=state["vehicles"],
        reference_time=state["reference_time"],
    )
    result = optimizer.optimize()
    state["result"] = result

    charts = generate_chart_data(result)
    return {
        "result": result.model_dump(mode="json"),
        "charts": charts,
    }


@app.get("/api/map", response_class=HTMLResponse)
async def route_map():
    if not state["result"]:
        return HTMLResponse("<h3>No optimization results yet.</h3>")
    return HTMLResponse(generate_route_map(state["result"], state["orders"]))


@app.post("/api/simulate")
async def run_simulation():
    """Run multiple scenarios and compare results."""
    scenarios = [
        {"name": "Light Load", "orders": 8, "drivers": 3, "vehicles": 3, "scenario": "normal"},
        {"name": "Normal", "orders": 15, "drivers": 3, "vehicles": 3, "scenario": "normal"},
        {"name": "Peak - Few Drivers", "orders": 20, "drivers": 2, "vehicles": 2, "scenario": "peak"},
        {"name": "Peak - Adequate", "orders": 20, "drivers": 4, "vehicles": 4, "scenario": "peak"},
        {"name": "Heavy Load", "orders": 35, "drivers": 5, "vehicles": 5, "scenario": "peak"},
    ]

    results = []
    ref_time = datetime.now().replace(
        hour=OPERATING_START_HOUR, minute=0, second=0, microsecond=0
    )

    for sc in scenarios:
        orders = generate_orders(sc["orders"], ref_time, sc["scenario"])
        drivers = generate_drivers(sc["drivers"], ref_time)
        vehicles = generate_vehicles(sc["vehicles"])

        optimizer = RouteOptimizer(orders, drivers, vehicles, ref_time)
        res = optimizer.optimize()

        results.append({
            "scenario": sc["name"],
            "total_orders": sc["orders"],
            "drivers": sc["drivers"],
            "vehicles": sc["vehicles"],
            "routes_created": len(res.routes),
            "assigned": res.assigned_orders,
            "unassigned": len(res.unassigned_orders),
            "deferred": len(res.deferred_orders),
            "sla_pct": res.sla_compliance_pct,
            "total_distance_km": res.total_distance_km,
            "total_duration_min": res.total_duration_min,
            "opt_time_sec": res.optimization_time_sec,
            "driver_util_pct": res.driver_utilization_pct,
        })

    return {"simulations": results}
